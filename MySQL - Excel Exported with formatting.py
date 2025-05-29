
import mysql.connector
import pandas as pd
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl import load_workbook

# connect to MySQL
conn=mysql.connector.connect(
host='localhost',
   user='root',
   password='Mysql@4444',
   database='northwind'  # or your schema name
 )
query = "select * from strings"
cursor = conn.cursor()
cursor.execute(query)
columns = [col[0] for col in cursor.description]
data = cursor.fetchall()
df = pd.DataFrame(data, columns=columns)

# save DataFrame to Excel
file_name = "String_.xlsx"
df.to_excel(file_name, index=False, engine='openpyxl')

# Apply formatting with openpyxl
wb = load_workbook(file_name)
ws = wb.active

# Styling
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="4F81BD")
alignment = Alignment(horizontal="center", vertical="center")
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
    )

#Apply header style
for cell in ws[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = alignment
    cell.border = thin_border

for row in ws.iter_rows(min_row=2,
                        max_row=ws.max_row,
                        max_col=ws.max_column):
    for cell in row:
        cell.alignment = alignment
        cell.border = thin_border

# Save styled workbook
wb.save(file_name)
print("Excel exported:", file_name)

cursor.close()
conn.close()
cursor.close()
conn.close()
print("Table is now successfully exported")
